from __future__ import annotations

import hashlib
import hmac
import secrets
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Iterable
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import bangla
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .db import get_connection

UTC = timezone.utc
try:
    IST = ZoneInfo("Asia/Kolkata")
except ZoneInfoNotFoundError:
    # Fallback for systems missing IANA tz database.
    IST = timezone(timedelta(hours=5, minutes=30))


def now_utc_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def utc_to_ist_string(value: str | None) -> str:
    if not value:
        return ""
    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S")


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 150_000)
    return f"{salt.hex()}:{digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        salt_hex, digest_hex = stored_hash.split(":", 1)
    except ValueError:
        return False
    salt = bytes.fromhex(salt_hex)
    expected = bytes.fromhex(digest_hex)
    test = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 150_000)
    return hmac.compare_digest(test, expected)


def create_user(username: str, password: str) -> None:
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
            (username.strip(), hash_password(password), now_utc_iso()),
        )
        conn.commit()


def list_users() -> list[str]:
    with get_connection() as conn:
        rows = conn.execute("SELECT username FROM users ORDER BY username COLLATE NOCASE").fetchall()
    return [row[0] for row in rows]


def update_user_password(username: str, new_password: str) -> None:
    stamp = now_utc_iso()
    hashed = hash_password(new_password)
    with get_connection() as conn:
        conn.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE username = ?",
            (hashed, stamp, username.strip()),
        )
        conn.commit()


def authenticate_user(username: str, password: str) -> bool:
    with get_connection() as conn:
        row = conn.execute(
            "SELECT password_hash FROM users WHERE username = ?",
            (username.strip(),),
        ).fetchone()
    if not row:
        return False
    return verify_password(password, row["password_hash"])


def list_clients(search: str = "") -> list[dict]:
    query = "SELECT * FROM clients WHERE deleted_at IS NULL"
    params: list[str] = []
    if search.strip():
        query += " AND (name LIKE ? OR phone LIKE ? OR email LIKE ?)"
        term = f"%{search.strip()}%"
        params.extend([term, term, term])
    query += " ORDER BY name COLLATE NOCASE"
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(row) for row in rows]


def get_client(client_id: int) -> dict | None:
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM clients WHERE id = ? AND deleted_at IS NULL",
            (client_id,),
        ).fetchone()
    return dict(row) if row else None


def create_client(name: str, phone: str, email: str, client_type: str = "buyer") -> None:
    stamp = now_utc_iso()
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO clients (name, phone, email, client_type, total_amount, amount_paid, created_at, updated_at)
            VALUES (?, ?, ?, ?, 0, 0, ?, ?)
            """,
            (name.strip(), phone.strip(), email.strip(), client_type.strip(), stamp, stamp),
        )
        conn.commit()


def update_client(client_id: int, name: str, phone: str, email: str, client_type: str = "buyer") -> None:
    with get_connection() as conn:
        conn.execute(
            """
            UPDATE clients
            SET name = ?, phone = ?, email = ?, client_type = ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (name.strip(), phone.strip(), email.strip(), client_type.strip(), now_utc_iso(), client_id),
        )
        conn.commit()


def delete_client(client_id: int) -> None:
    stamp = now_utc_iso()
    with get_connection() as conn:
        conn.execute(
            "UPDATE clients SET deleted_at = ?, updated_at = ? WHERE id = ? AND deleted_at IS NULL",
            (stamp, stamp, client_id),
        )
        conn.execute(
            "UPDATE ledgers SET deleted_at = ?, updated_at = ? WHERE client_id = ? AND deleted_at IS NULL",
            (stamp, stamp, client_id),
        )
        conn.commit()


def list_ledgers(client_id: int) -> list[dict]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM ledgers
            WHERE client_id = ? AND deleted_at IS NULL
            ORDER BY created_at DESC, id DESC
            """,
            (client_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def list_ledgers_client_wise(client_id: int | None = None) -> list[dict]:
    query = """
        SELECT l.id, l.client_id, c.name AS client_name, l.product_name, l.quantity_kg, l.price_per_kg,
               l.total_price, l.created_at, l.updated_at
        FROM ledgers l
        JOIN clients c ON c.id = l.client_id
        WHERE l.deleted_at IS NULL AND c.deleted_at IS NULL
    """
    params: list[int] = []
    if client_id is not None:
        query += " AND l.client_id = ?"
        params.append(client_id)
    query += " ORDER BY l.created_at DESC, l.id DESC"
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(row) for row in rows]


def create_ledgers(client_id: int, entries: Iterable[dict[str, float | str]]) -> list[dict]:
    entries = list(entries)
    if not entries:
        return []

    stamp = now_utc_iso()
    rows: list[dict] = []
    total_added = 0.0

    with get_connection() as conn:
        for entry in entries:
            qty = float(entry["quantity_kg"])
            price = float(entry["price_per_kg"])
            total_price = qty * price
            total_added += total_price
            product_name = str(entry.get("product_name", "")).strip()
            entry_created_at = str(entry.get("created_at", stamp)).strip()
            if not entry_created_at:
                entry_created_at = stamp
            
            cursor = conn.execute(
                """
                INSERT INTO ledgers (client_id, product_name, quantity_kg, price_per_kg, total_price, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (client_id, product_name, qty, price, total_price, entry_created_at, stamp),
            )
            rows.append(
                {
                    "id": cursor.lastrowid,
                    "client_id": client_id,
                    "product_name": product_name,
                    "quantity_kg": qty,
                    "price_per_kg": price,
                    "total_price": total_price,
                    "created_at": entry_created_at,
                    "updated_at": stamp,
                }
            )

        conn.execute(
            """
            UPDATE clients
            SET total_amount = COALESCE(total_amount, 0) + ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (total_added, stamp, client_id),
        )
        conn.commit()

    return rows


def delete_ledger(ledger_id: int) -> None:
    stamp = now_utc_iso()
    with get_connection() as conn:
        conn.execute(
            "UPDATE ledgers SET deleted_at = ?, updated_at = ? WHERE id = ? AND deleted_at IS NULL",
            (stamp, stamp, ledger_id),
        )
        conn.commit()


def export_client_ledgers(client_id: int) -> BytesIO | None:
    client = get_client(client_id)
    if not client:
        return None

    ledgers = list_ledgers(client_id)
    payments = list_payments(client_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    # Determine layout based on client type
    is_buyer = client.get("client_type", "buyer").lower() == "buyer"
    
    if is_buyer:
        # For Buyer: Payments on left, Products on right
        # Payments section
        sheet["A1"] = "PAYMENT RECORDS"
        sheet["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        sheet["A1"].fill = PatternFill("solid", fgColor="00B050")
        
        payment_headers = ["ID", "Amount", "Mode", "Date"]
        for col, header in enumerate(payment_headers, start=1):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        
        for idx, payment in enumerate(payments, start=3):
            sheet.cell(row=idx, column=1, value=payment["id"])
            sheet.cell(row=idx, column=2, value=payment["amount"])
            sheet.cell(row=idx, column=3, value=payment["payment_mode"])
            sheet.cell(row=idx, column=4, value=utc_to_bengali_date(payment["created_at"]))
        
        # Products section
        sheet["E1"] = "PRODUCT RECORDS"
        sheet["E1"].font = Font(bold=True, color="FFFFFF", size=12)
        sheet["E1"].fill = PatternFill("solid", fgColor="0070C0")
        
        product_headers = ["ID", "Product", "Qty (kg)", "Price/kg", "Total", "Date"]
        for col, header in enumerate(product_headers, start=5):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        
        for idx, ledger in enumerate(ledgers, start=3):
            sheet.cell(row=idx, column=5, value=ledger["id"])
            sheet.cell(row=idx, column=6, value=ledger.get("product_name", ""))
            sheet.cell(row=idx, column=7, value=ledger["quantity_kg"])
            sheet.cell(row=idx, column=8, value=ledger["price_per_kg"])
            sheet.cell(row=idx, column=9, value=ledger["total_price"])
            sheet.cell(row=idx, column=10, value=utc_to_bengali_date(ledger["created_at"]))
    else:
        # For Seller: Products on left, Payments on right
        # Products section
        sheet["A1"] = "PRODUCT RECORDS"
        sheet["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        sheet["A1"].fill = PatternFill("solid", fgColor="0070C0")
        
        product_headers = ["ID", "Product", "Qty (kg)", "Price/kg", "Total", "Date"]
        for col, header in enumerate(product_headers, start=1):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        
        for idx, ledger in enumerate(ledgers, start=3):
            sheet.cell(row=idx, column=1, value=ledger["id"])
            sheet.cell(row=idx, column=2, value=ledger.get("product_name", ""))
            sheet.cell(row=idx, column=3, value=ledger["quantity_kg"])
            sheet.cell(row=idx, column=4, value=ledger["price_per_kg"])
            sheet.cell(row=idx, column=5, value=ledger["total_price"])
            sheet.cell(row=idx, column=6, value=utc_to_bengali_date(ledger["created_at"]))
        
        # Payments section
        sheet["H1"] = "PAYMENT RECORDS"
        sheet["H1"].font = Font(bold=True, color="FFFFFF", size=12)
        sheet["H1"].fill = PatternFill("solid", fgColor="00B050")
        
        payment_headers = ["ID", "Amount", "Mode", "Date"]
        for col, header in enumerate(payment_headers, start=8):
            cell = sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        
        for idx, payment in enumerate(payments, start=3):
            sheet.cell(row=idx, column=8, value=payment["id"])
            sheet.cell(row=idx, column=9, value=payment["amount"])
            sheet.cell(row=idx, column=10, value=payment["payment_mode"])
            sheet.cell(row=idx, column=11, value=utc_to_bengali_date(payment["created_at"]))

    # Summary sheet
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Client Name"
    summary["B1"] = client["name"]
    summary["A2"] = "Client Type"
    summary["B2"] = client.get("client_type", "buyer").capitalize()
    summary["A3"] = "Phone"
    summary["B3"] = client["phone"]
    summary["A4"] = "Email"
    summary["B4"] = client["email"]
    summary["A5"] = "Total Amount"
    summary["B5"] = client["total_amount"]
    summary["A6"] = "Amount Paid"
    summary["B6"] = client.get("amount_paid", 0)
    summary["A7"] = "Remaining Amount"
    summary["B7"] = client["total_amount"] - client.get("amount_paid", 0)
    summary["A8"] = "Exported At"
    summary["B8"] = utc_to_bengali_date(datetime.now(IST).isoformat())

    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 40

    # Auto-adjust column widths for data sheet
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        if length > 0:
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 22)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer

def clear_client_ledgers(client_id: int) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM ledgers WHERE client_id = ?", (client_id,))
        conn.commit()


def clear_database() -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM ledgers")
        conn.execute("DELETE FROM clients")
        conn.execute("DELETE FROM users")
        conn.execute("DELETE FROM payments")
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('ledgers', 'clients', 'users', 'payments')")
        conn.commit()


def utc_to_bengali_date(value: str | None) -> str:
    """Convert a UTC ISO timestamp to a Bengali calendar date using bangla."""
    if not value:
        return ""

    dt = datetime.fromisoformat(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    ist_dt = dt.astimezone(IST)

    bangla_date = bangla.get_date(ist_dt.day, ist_dt.month, ist_dt.year)
    date = bangla_date.get("date", "")
    month = bangla_date.get("month", "")
    year = bangla_date.get("year", "")
    return f"{date} {month} {year}".strip()


# Payment Management Functions

def create_payment(client_id: int, amount: float, payment_mode: str, created_at: str | None = None) -> None:
    """Create a new payment record for a client."""
    stamp = now_utc_iso()
    if created_at is None:
        created_at = stamp
    
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO payments (client_id, amount, payment_mode, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (client_id, amount, payment_mode.strip(), created_at, stamp),
        )
        
        # Update client amount_paid
        conn.execute(
            """
            UPDATE clients
            SET amount_paid = COALESCE(amount_paid, 0) + ?, updated_at = ?
            WHERE id = ? AND deleted_at IS NULL
            """,
            (amount, stamp, client_id),
        )
        conn.commit()


def list_payments(client_id: int | None = None) -> list[dict]:
    """List all payment records, optionally filtered by client."""
    query = """
        SELECT p.id, p.client_id, c.name AS client_name, p.amount, p.payment_mode, 
               p.created_at, p.updated_at
        FROM payments p
        JOIN clients c ON c.id = p.client_id
        WHERE p.deleted_at IS NULL AND c.deleted_at IS NULL
    """
    params: list[int] = []
    if client_id is not None:
        query += " AND p.client_id = ?"
        params.append(client_id)
    query += " ORDER BY p.created_at DESC, p.id DESC"
    
    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(row) for row in rows]


def delete_payment(payment_id: int) -> None:
    """Soft delete a payment record."""
    stamp = now_utc_iso()
    
    with get_connection() as conn:
        # Get payment amount before deleting
        payment = conn.execute(
            "SELECT amount, client_id FROM payments WHERE id = ? AND deleted_at IS NULL",
            (payment_id,),
        ).fetchone()
        
        if payment:
            # Update payment as deleted
            conn.execute(
                "UPDATE payments SET deleted_at = ?, updated_at = ? WHERE id = ? AND deleted_at IS NULL",
                (stamp, stamp, payment_id),
            )
            
            # Reduce client amount_paid
            conn.execute(
                """
                UPDATE clients
                SET amount_paid = COALESCE(amount_paid, 0) - ?, updated_at = ?
                WHERE id = ? AND deleted_at IS NULL
                """,
                (payment["amount"], stamp, payment["client_id"]),
            )
        
        conn.commit()


def get_client_summary(client_id: int) -> dict | None:
    """Get summary info for a client including payments."""
    client = get_client(client_id)
    if not client:
        return None
    
    with get_connection() as conn:
        # Get total payment
        payment_row = conn.execute(
            """
            SELECT COALESCE(SUM(amount), 0) as total_paid
            FROM payments
            WHERE client_id = ? AND deleted_at IS NULL
            """,
            (client_id,),
        ).fetchone()
        
        total_paid = payment_row["total_paid"] if payment_row else 0
    
    remaining = client["total_amount"] - total_paid
    
    return {
        **client,
        "total_paid": total_paid,
        "remaining_amount": remaining
    }
